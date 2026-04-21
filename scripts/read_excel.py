import openpyxl, json, sys
wb = openpyxl.load_workbook(r"N:\商品資料庫AI專用-開放\商品優化原文件\澳思萊_產品資料庫_162款_完整版.xlsx", data_only=True)
print("SHEETS:", wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"--- {s} ({ws.max_row} rows x {ws.max_column} cols) ---")
    # print first 3 rows to see structure
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        print(i, row)
        if i >= 2: break
